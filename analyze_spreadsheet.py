#!/usr/bin/env python
"""
Script to analyze the Excel spreadsheet structure for ISO/Contract-Projects-Planned Audits 2
and Contacts sheets.
"""
import openpyxl
import json
import sys
import os

def analyze_sheet(sheet):
    """Analyze a worksheet and extract its structure."""
    result = {
        'name': sheet.title,
        'dimensions': f"{sheet.max_row} rows x {sheet.max_column} columns",
        'max_row': sheet.max_row,
        'max_column': sheet.max_column,
        'headers': [],
        'sample_data': [],
        'column_details': []
    }
    
    # Get headers (first row)
    if sheet.max_row > 0:
        headers = []
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col)
            header = cell.value if cell.value else f"Column_{col}"
            headers.append(header)
        result['headers'] = headers
        
        # Get sample data (first 5 rows after header)
        sample_rows = []
        for row in range(2, min(7, sheet.max_row + 1)):
            row_data = {}
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                header = headers[col - 1]
                row_data[header] = str(cell.value) if cell.value is not None else ""
            sample_rows.append(row_data)
        result['sample_data'] = sample_rows
        
        # Analyze column details
        for col_idx, header in enumerate(headers, start=1):
            col_info = {
                'index': col_idx,
                'name': header,
                'sample_values': [],
                'non_empty_count': 0,
                'data_types': set()
            }
            
            # Sample values from column
            for row in range(2, min(12, sheet.max_row + 1)):
                cell = sheet.cell(row=row, column=col_idx)
                if cell.value is not None:
                    col_info['non_empty_count'] += 1
                    col_info['sample_values'].append(str(cell.value)[:100])
                    col_info['data_types'].add(type(cell.value).__name__)
            
            col_info['data_types'] = list(col_info['data_types'])
            result['column_details'].append(col_info)
    
    return result

def main():
    # Path to the Excel file
    excel_file = '../iso/Contacts - Projects - Planned Audits 2.xlsx'
    
    if not os.path.exists(excel_file):
        print(f"Error: File not found: {excel_file}")
        sys.exit(1)
    
    print(f"Analyzing: {excel_file}")
    print("=" * 80)
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    print(f"\nAvailable sheets: {wb.sheetnames}\n")
    
    analysis_results = {}
    
    # Analyze each sheet
    for sheet_name in wb.sheetnames:
        print(f"\n{'=' * 80}")
        print(f"Analyzing sheet: {sheet_name}")
        print('=' * 80)
        
        sheet = wb[sheet_name]
        result = analyze_sheet(sheet)
        analysis_results[sheet_name] = result
        
        print(f"\nDimensions: {result['dimensions']}")
        print(f"\nHeaders ({len(result['headers'])}):")
        for idx, header in enumerate(result['headers'], 1):
            print(f"  {idx:2d}. {header}")
        
        print(f"\nColumn Details:")
        for col in result['column_details']:
            print(f"\n  Column {col['index']}: {col['name']}")
            print(f"    - Non-empty cells: {col['non_empty_count']}")
            print(f"    - Data types: {', '.join(col['data_types'])}")
            if col['sample_values']:
                print(f"    - Sample values: {col['sample_values'][:3]}")
    
    # Save to JSON
    output_file = 'spreadsheet_analysis.json'
    with open(output_file, 'w') as f:
        json.dump(analysis_results, f, indent=2, default=str)
    
    print(f"\n{'=' * 80}")
    print(f"Analysis saved to: {output_file}")
    print('=' * 80)

if __name__ == '__main__':
    main()

