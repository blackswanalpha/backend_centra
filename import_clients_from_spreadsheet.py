#!/usr/bin/env python
"""
Script to import client data from Excel spreadsheet into the Clients database.
Reads from "Contacts - Projects - Planned Audits 2.xlsx" and populates the Client model.
"""
import os
import sys
import django
from datetime import datetime
import openpyxl

# Setup Django environment
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'backend.settings')
django.setup()

from apps.clients.models import Client
from django.db import transaction


class ClientImporter:
    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path
        self.workbook = None
        self.contacts_sheet = None
        self.projects_sheet = None
        self.stats = {
            'total_contacts': 0,
            'total_projects': 0,
            'clients_created': 0,
            'clients_updated': 0,
            'errors': []
        }
    
    def load_workbook(self):
        """Load the Excel workbook."""
        print(f"Loading workbook: {self.excel_file_path}")
        self.workbook = openpyxl.load_workbook(self.excel_file_path, data_only=True)
        self.contacts_sheet = self.workbook['Contacts']
        self.projects_sheet = self.workbook['Project']
        print(f"✓ Loaded sheets: {self.workbook.sheetnames}")
    
    def parse_contacts_sheet(self):
        """Parse the Contacts sheet and return a dictionary of client data."""
        contacts_data = {}
        
        # Skip header row
        for row_idx, row in enumerate(self.contacts_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # Skip empty rows
                continue
            
            client_name = row[0].strip() if row[0] else None
            if not client_name:
                continue
            
            # Remove currency suffix from client name (e.g., "Client (USD)" -> "Client")
            clean_name = client_name
            if '(' in client_name and ')' in client_name:
                clean_name = client_name.split('(')[0].strip()
            
            contacts_data[clean_name] = {
                'name': clean_name,
                'currency_code': row[1] if row[1] else None,
                'billing_attention': row[2] if row[2] else None,
                'billing_address': row[3] if row[3] else None,
                'billing_street2': row[4] if row[4] else None,
                'billing_city': row[5] if row[5] else None,
                'billing_state': row[6] if row[6] else None,
                'billing_country': row[7] if row[7] else None,
                'payment_terms': row[8] if row[8] else None,
            }
            self.stats['total_contacts'] += 1
        
        print(f"✓ Parsed {len(contacts_data)} unique clients from Contacts sheet")
        return contacts_data
    
    def parse_projects_sheet(self):
        """Parse the Projects sheet and return a list of project data."""
        projects_data = []
        
        # Skip header row
        for row_idx, row in enumerate(self.projects_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[1]:  # Skip if no client name
                continue
            
            client_name = row[1].strip() if row[1] else None
            if not client_name:
                continue
            
            # Parse dates
            registration_date = row[9] if isinstance(row[9], datetime) else None
            certificate_date = row[10] if isinstance(row[10], datetime) else None
            expiry_date = row[11] if isinstance(row[11], datetime) else None
            
            project_data = {
                'project_ref': row[0] if row[0] else None,
                'client_name': client_name,
                'physical_address': row[2] if row[2] else None,
                'city': row[3] if row[3] else None,
                'country': row[4] if row[4] else None,
                'standard': row[5] if row[5] else None,
                'description': row[6] if row[6] else None,
                'scope_of_certification': row[7] if row[7] else None,
                'certificate_no': row[8] if row[8] else None,
                'registration_date': registration_date.date() if registration_date else None,
                'certificate_date': certificate_date.date() if certificate_date else None,
                'expiry_date': expiry_date.date() if expiry_date else None,
            }
            projects_data.append(project_data)
            self.stats['total_projects'] += 1
        
        print(f"✓ Parsed {len(projects_data)} projects from Projects sheet")
        return projects_data
    
    def merge_client_data(self, contacts_data, projects_data):
        """Merge contacts and projects data by client name."""
        merged_data = {}
        
        # Start with contacts data
        for client_name, contact_info in contacts_data.items():
            merged_data[client_name] = contact_info.copy()
            merged_data[client_name]['projects'] = []
        
        # Add projects to corresponding clients
        for project in projects_data:
            client_name = project['client_name']
            
            # Try to find matching client (exact match or partial match)
            matched_client = None
            if client_name in merged_data:
                matched_client = client_name
            else:
                # Try partial match
                for existing_name in merged_data.keys():
                    if existing_name.lower() in client_name.lower() or client_name.lower() in existing_name.lower():
                        matched_client = existing_name
                        break
            
            if matched_client:
                merged_data[matched_client]['projects'].append(project)
            else:
                # Create new client entry from project data
                if client_name not in merged_data:
                    merged_data[client_name] = {
                        'name': client_name,
                        'projects': [project]
                    }
        
        return merged_data
    
    def import_clients(self, merged_data, dry_run=True):
        """Import or update clients in the database."""
        print(f"\n{'DRY RUN - ' if dry_run else ''}Importing clients...")
        
        for client_name, client_data in merged_data.items():
            try:
                # Check if client exists
                existing_client = Client.objects.filter(name__iexact=client_name).first()
                
                if existing_client:
                    # Update existing client
                    self.update_client(existing_client, client_data, dry_run)
                    self.stats['clients_updated'] += 1
                else:
                    # Create new client
                    self.create_client(client_data, dry_run)
                    self.stats['clients_created'] += 1
            
            except Exception as e:
                error_msg = f"Error processing client '{client_name}': {str(e)}"
                self.stats['errors'].append(error_msg)
                print(f"  ✗ {error_msg}")

        return self.stats

    def create_client(self, client_data, dry_run=True):
        """Create a new client."""
        # Get first project data if available
        first_project = client_data.get('projects', [{}])[0] if client_data.get('projects') else {}

        # Collect all certifications from projects
        certifications = []
        for project in client_data.get('projects', []):
            if project.get('standard') and project['standard'] not in certifications:
                certifications.append(project['standard'])

        # Determine address - use billing address, physical address, or default
        address = (client_data.get('billing_address') or
                   first_project.get('physical_address') or
                   'N/A')

        client_fields = {
            'name': client_data.get('name'),
            'address': address,

            # Billing information
            'currency_code': client_data.get('currency_code'),
            'billing_attention': client_data.get('billing_attention'),
            'billing_address': client_data.get('billing_address'),
            'billing_street2': client_data.get('billing_street2'),
            'billing_city': client_data.get('billing_city'),
            'billing_state': client_data.get('billing_state'),
            'billing_country': client_data.get('billing_country'),
            'payment_terms': client_data.get('payment_terms'),

            # Physical location (from first project)
            'physical_address': first_project.get('physical_address'),
            'city': first_project.get('city'),
            'country': first_project.get('country'),

            # Project/Certification info (from first project)
            'project_ref': first_project.get('project_ref'),
            'scope_of_certification': first_project.get('scope_of_certification'),
            'certificate_no': first_project.get('certificate_no'),
            'registration_date': first_project.get('registration_date'),
            'certificate_date': first_project.get('certificate_date'),
            'expiry_date': first_project.get('expiry_date'),

            # Certifications array
            'certifications': certifications,

            # Default values for required fields
            'contact': client_data.get('billing_attention') or 'N/A',
            'email': 'noemail@example.com',  # Placeholder
            'phone': 'N/A',  # Placeholder
        }

        if not dry_run:
            client = Client.objects.create(**client_fields)
            print(f"  ✓ Created client: {client.name} (ID: {client.id})")
        else:
            print(f"  [DRY RUN] Would create client: {client_data.get('name')}")

    def update_client(self, client, client_data, dry_run=True):
        """Update an existing client."""
        # Get first project data if available
        first_project = client_data.get('projects', [{}])[0] if client_data.get('projects') else {}

        # Collect all certifications from projects
        certifications = list(client.certifications) if client.certifications else []
        for project in client_data.get('projects', []):
            if project.get('standard') and project['standard'] not in certifications:
                certifications.append(project['standard'])

        # Update fields
        updates = []

        # Billing information
        if client_data.get('currency_code') and not client.currency_code:
            client.currency_code = client_data['currency_code']
            updates.append('currency_code')

        if client_data.get('billing_attention') and not client.billing_attention:
            client.billing_attention = client_data['billing_attention']
            updates.append('billing_attention')

        if client_data.get('billing_address') and not client.billing_address:
            client.billing_address = client_data['billing_address']
            updates.append('billing_address')

        if client_data.get('billing_street2') and not client.billing_street2:
            client.billing_street2 = client_data['billing_street2']
            updates.append('billing_street2')

        if client_data.get('billing_city') and not client.billing_city:
            client.billing_city = client_data['billing_city']
            updates.append('billing_city')

        if client_data.get('billing_state') and not client.billing_state:
            client.billing_state = client_data['billing_state']
            updates.append('billing_state')

        if client_data.get('billing_country') and not client.billing_country:
            client.billing_country = client_data['billing_country']
            updates.append('billing_country')

        if client_data.get('payment_terms') and not client.payment_terms:
            client.payment_terms = client_data['payment_terms']
            updates.append('payment_terms')

        # Physical location
        if first_project.get('physical_address') and not client.physical_address:
            client.physical_address = first_project['physical_address']
            updates.append('physical_address')

        if first_project.get('city') and not client.city:
            client.city = first_project['city']
            updates.append('city')

        if first_project.get('country') and not client.country:
            client.country = first_project['country']
            updates.append('country')

        # Project/Certification info
        if first_project.get('project_ref') and not client.project_ref:
            client.project_ref = first_project['project_ref']
            updates.append('project_ref')

        if first_project.get('scope_of_certification') and not client.scope_of_certification:
            client.scope_of_certification = first_project['scope_of_certification']
            updates.append('scope_of_certification')

        if first_project.get('certificate_no') and not client.certificate_no:
            client.certificate_no = first_project['certificate_no']
            updates.append('certificate_no')

        if first_project.get('registration_date') and not client.registration_date:
            client.registration_date = first_project['registration_date']
            updates.append('registration_date')

        if first_project.get('certificate_date') and not client.certificate_date:
            client.certificate_date = first_project['certificate_date']
            updates.append('certificate_date')

        if first_project.get('expiry_date') and not client.expiry_date:
            client.expiry_date = first_project['expiry_date']
            updates.append('expiry_date')

        # Update certifications
        if certifications != client.certifications:
            client.certifications = certifications
            updates.append('certifications')

        if updates:
            if not dry_run:
                client.save()
                print(f"  ✓ Updated client: {client.name} (Fields: {', '.join(updates)})")
            else:
                print(f"  [DRY RUN] Would update client: {client.name} (Fields: {', '.join(updates)})")
        else:
            print(f"  - No updates needed for: {client.name}")

    def run(self, dry_run=True):
        """Run the full import process."""
        print("=" * 80)
        print("CLIENT DATA IMPORT FROM SPREADSHEET")
        print("=" * 80)

        self.load_workbook()

        # Parse sheets
        contacts_data = self.parse_contacts_sheet()
        projects_data = self.parse_projects_sheet()

        # Merge data
        merged_data = self.merge_client_data(contacts_data, projects_data)
        print(f"✓ Merged data for {len(merged_data)} unique clients")

        # Import clients
        with transaction.atomic():
            stats = self.import_clients(merged_data, dry_run=dry_run)

            if dry_run:
                print("\n" + "=" * 80)
                print("DRY RUN COMPLETE - No changes made to database")
                print("=" * 80)
            else:
                print("\n" + "=" * 80)
                print("IMPORT COMPLETE")
                print("=" * 80)

        # Print statistics
        print(f"\nStatistics:")
        print(f"  - Total contacts parsed: {stats['total_contacts']}")
        print(f"  - Total projects parsed: {stats['total_projects']}")
        print(f"  - Clients created: {stats['clients_created']}")
        print(f"  - Clients updated: {stats['clients_updated']}")
        print(f"  - Errors: {len(stats['errors'])}")

        if stats['errors']:
            print("\nErrors:")
            for error in stats['errors']:
                print(f"  - {error}")

        return stats


def main():
    import argparse

    parser = argparse.ArgumentParser(description='Import client data from Excel spreadsheet')
    parser.add_argument('--file', default='../iso/Contacts - Projects - Planned Audits 2.xlsx',
                        help='Path to Excel file')
    parser.add_argument('--execute', action='store_true',
                        help='Execute the import (default is dry run)')

    args = parser.parse_args()

    importer = ClientImporter(args.file)
    importer.run(dry_run=not args.execute)


if __name__ == '__main__':
    main()


